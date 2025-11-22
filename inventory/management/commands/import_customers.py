from __future__ import annotations

import unicodedata
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Q

from inventory.models import Customer, generate_customer_reference

try:  # pragma: no cover
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]


NAME_HEADERS = ("nom complet", "nom", "name")
PHONE_HEADERS = ("telephone", "tel", "phone")
EMAIL_HEADERS = ("email", "mail")
VENDOR_HEADERS = ("vendeur", "salesperson")
ACTIVITY_HEADERS = ("activite", "activites", "activity")
CITY_HEADERS = ("ville", "city")
COUNTRY_HEADERS = ("pays", "country")


def _normalize_header(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def _clean_text(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def _select_column(headers: tuple[object, ...], candidates: tuple[str, ...]) -> int | None:
    normalized_headers = [_normalize_header(header) for header in headers]
    for candidate in candidates:
        for idx, header in enumerate(normalized_headers):
            if candidate in header:
                return idx
    return None


def _pick_value(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None:
        return None
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _find_existing_customer(name: str, email: str, phone: str) -> Customer | None:
    filters = Q(name__iexact=name) | Q(company_name__iexact=name)
    if email:
        filters |= Q(email__iexact=email)
    if phone:
        filters |= Q(phone__iexact=phone)
    return Customer.objects.filter(filters).first()


class Command(BaseCommand):
    help = "Importe des clients depuis un fichier Excel (.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument(
            "file",
            type=str,
            help="Chemin du fichier Excel contenant les clients.",
        )
        parser.add_argument(
            "--sheet",
            "-s",
            type=str,
            default=None,
            help="Nom ou index (0-based) de la feuille a utiliser (defaut: feuille active).",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError("openpyxl est requis pour lire les fichiers Excel (pip install openpyxl).")

        file_path = Path(options["file"]).expanduser()
        if not file_path.exists():
            raise CommandError(f"Le fichier '{file_path}' est introuvable.")
        if not file_path.is_file():
            raise CommandError(f"'{file_path}' n'est pas un fichier.")

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet_identifier = options.get("sheet")
        sheet = self._select_sheet(workbook, sheet_identifier)
        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headers:
            raise CommandError("Impossible de lire l'en-tete du fichier Excel.")

        name_idx = _select_column(headers, NAME_HEADERS)
        if name_idx is None:
            raise CommandError("Colonne contenant le nom introuvable (ex: 'Nom complet').")

        phone_idx = _select_column(headers, PHONE_HEADERS)
        email_idx = _select_column(headers, EMAIL_HEADERS)
        vendor_idx = _select_column(headers, VENDOR_HEADERS)
        activity_idx = _select_column(headers, ACTIVITY_HEADERS)
        city_idx = _select_column(headers, CITY_HEADERS)
        country_idx = _select_column(headers, COUNTRY_HEADERS)

        summary = {
            "rows": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
        }

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            summary["rows"] += 1
            name = _clean_text(_pick_value(row, name_idx))
            if not name:
                summary["skipped"] += 1
                continue

            phone = _clean_text(_pick_value(row, phone_idx))
            email = _clean_text(_pick_value(row, email_idx)).lower()
            vendor = _clean_text(_pick_value(row, vendor_idx))
            activity = _clean_text(_pick_value(row, activity_idx))
            city = _clean_text(_pick_value(row, city_idx))
            country = _clean_text(_pick_value(row, country_idx))

            address_parts = [part for part in (city, country) if part]
            address = ", ".join(address_parts)
            notes_parts = []
            if vendor:
                notes_parts.append(f"Vendeur: {vendor}")
            if activity:
                notes_parts.append(f"Activite: {activity}")
            notes = " | ".join(notes_parts)

            customer = _find_existing_customer(name, email, phone)
            created = False
            if customer is None:
                customer = Customer(
                    name=name,
                    reference=generate_customer_reference(),
                )
                created = True

            updated = False
            if phone and customer.phone != phone:
                customer.phone = phone
                updated = True
            if email and customer.email != email:
                customer.email = email
                updated = True
            if address and (not customer.address or customer.address != address):
                customer.address = address
                updated = True
            if notes:
                if not customer.notes:
                    customer.notes = notes
                    updated = True
                elif notes not in customer.notes:
                    customer.notes = f"{customer.notes} | {notes}"
                    updated = True

            if created or updated:
                customer.save()
                summary["created" if created else "updated"] += 1

        self.stdout.write(
            self.style.SUCCESS(
                "Import termine : %(created)d crees, %(updated)d mis a jour, %(skipped)d ignores (nom vide) sur %(rows)d lignes."
                % summary
            )
        )

    def _select_sheet(self, workbook, identifier: str | None):
        if identifier is None:
            return workbook.active
        cleaned = identifier.strip()
        if cleaned.isdigit():
            index = int(cleaned)
            try:
                return workbook.worksheets[index]
            except IndexError:
                raise CommandError(f"Feuille d'index {index} introuvable.")
        if cleaned in workbook.sheetnames:
            return workbook[cleaned]
        raise CommandError(f"Feuille '{identifier}' introuvable.")
