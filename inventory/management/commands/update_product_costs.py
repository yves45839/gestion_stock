from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence

from django.core.management.base import BaseCommand, CommandError

from inventory.models import Product

try:  # pragma: no cover
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]


DEFAULT_REFERENCE_HEADERS = (
    "référence interne",
    "reference interne",
    "reference",
    "ref",
    "sku",
    "code",
)
DEFAULT_COST_HEADERS = (
    "coût",
    "cout",
    "purchase_price",
    "prix achat",
    "prix de revient",
    "cost",
)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _column_letter_to_index(letter: str) -> int | None:
    cleaned = letter.strip().upper()
    if not cleaned.isalpha():
        return None
    index = 0
    for char in cleaned:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _choose_column_index(
    headers: Sequence[object], identifier: str | None, defaults: Iterable[str]
) -> int:
    if identifier:
        parsed = _parse_column_identifier(identifier, headers)
        if parsed is not None:
            return parsed
        raise CommandError(f"Impossible de localiser la colonne '{identifier}'.")
    for candidate in defaults:
        match = _find_header_index(headers, candidate)
        if match is not None:
            return match
    raise CommandError("Impossible de détecter la colonne par défaut (référence ou coût).")


def _parse_column_identifier(identifier: str, headers: Sequence[object]) -> int | None:
    raw = identifier.strip()
    if not raw:
        return None
    if raw.isdigit():
        idx = int(raw) - 1
        return idx if idx >= 0 else None
    letter_idx = _column_letter_to_index(raw)
    if letter_idx is not None:
        return letter_idx
    normalized = raw.lower()
    for idx, header in enumerate(headers):
        header_value = _normalize_header(header)
        if normalized in header_value:
            return idx
    return None


def _find_header_index(headers: Sequence[object], keyword: str) -> int | None:
    normalized_keyword = keyword.lower()
    for idx, header in enumerate(headers):
        header_value = _normalize_header(header)
        if normalized_keyword in header_value:
            return idx
    return None


def _select_sheet(workbook, sheet_identifier: str | None):
    if sheet_identifier is None:
        return workbook.active
    cleaned = sheet_identifier.strip()
    if cleaned.isdigit():
        index = int(cleaned)
        try:
            return workbook.worksheets[index]
        except IndexError:
            raise CommandError(f"Feuille d'index {index} introuvable.")
    if cleaned in workbook.sheetnames:
        return workbook[cleaned]
    raise CommandError(f"Feuille '{sheet_identifier}' introuvable.")


def _get_cell_value(row: Sequence[object], index: int) -> object | None:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _normalize_reference(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_decimal_value(raw_value: object | None) -> Decimal | None:
    if raw_value is None:
        return None
    if isinstance(raw_value, Decimal):
        return raw_value
    if isinstance(raw_value, (int, float)):
        try:
            return Decimal(str(raw_value))
        except InvalidOperation:
            return None
    text = str(raw_value).strip()
    if not text:
        return None
    text = re.sub(r"[^\d\-,\.]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "")
        text = text.replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    elif "." in text:
        parts = text.split(".")
        if len(parts) > 1 and len(parts[-1]) == 3:
            text = "".join(parts)
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


class Command(BaseCommand):
    help = "Met à jour le prix d'achat des produits à partir d'un fichier Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "file",
            type=str,
            help="Chemin vers le fichier Excel (.xlsx) contenant les références et les coûts.",
        )
        parser.add_argument(
            "--sheet",
            "-s",
            type=str,
            default="0",
            help="Nom ou index (0-based) de la feuille à utiliser (défaut: première feuille).",
        )
        parser.add_argument(
            "--reference-column",
            "-r",
            type=str,
            default=None,
            help="Colonne contenant la référence (nom de l'en-tête, lettre ou indice 1-based).",
        )
        parser.add_argument(
            "--cost-column",
            "-c",
            type=str,
            default=None,
            help="Colonne contenant le coût (nom de l'en-tête, lettre ou indice 1-based).",
        )
        parser.add_argument(
            "--match-field",
            "-m",
            choices=("manufacturer_reference", "sku"),
            default="manufacturer_reference",
            help="Champ produit utilisé pour retrouver la référence (défaut: manufacturer_reference).",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError("openpyxl est requis pour traiter les fichiers Excel (pip install openpyxl).")

        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"Le fichier '{file_path}' est introuvable.")
        if not file_path.is_file():
            raise CommandError(f"'{file_path}' n'est pas un fichier.")

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet = _select_sheet(workbook, options.get("sheet"))
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if not header_row:
            raise CommandError("Impossible de lire l'en-tête du fichier Excel.")

        reference_idx = _choose_column_index(
            header_row, options.get("reference_column"), DEFAULT_REFERENCE_HEADERS
        )
        cost_idx = _choose_column_index(
            header_row, options.get("cost_column"), DEFAULT_COST_HEADERS
        )

        summary = {
            "rows": 0,
            "updated": 0,
            "missing_reference": 0,
            "invalid_cost": 0,
            "not_found": 0,
            "ambiguous": 0,
        }
        errors: list[str] = []
        match_field = options["match_field"]

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            summary["rows"] += 1
            reference_value = _normalize_reference(_get_cell_value(row, reference_idx))
            if not reference_value:
                summary["missing_reference"] += 1
                errors.append(f"Ligne {row_number}: référence manquante.")
                continue

            cost_value = _parse_decimal_value(_get_cell_value(row, cost_idx))
            if cost_value is None:
                raw = _get_cell_value(row, cost_idx)
                summary["invalid_cost"] += 1
                errors.append(f"Ligne {row_number}: coût invalide '{raw}'.")
                continue

            lookup = {f"{match_field}__iexact": reference_value}
            products = Product.objects.filter(**lookup)
            count = products.count()
            if count == 0:
                summary["not_found"] += 1
                errors.append(f"Ligne {row_number}: référence '{reference_value}' introuvable.")
                continue
            if count > 1:
                summary["ambiguous"] += 1
                errors.append(
                    f"Ligne {row_number}: plusieurs produits correspondent à '{reference_value}'."
                )
                continue

            product = products.first()
            if product.purchase_price == cost_value:
                continue
            product.purchase_price = cost_value
            product.save(update_fields=["purchase_price", "updated_at"])
            summary["updated"] += 1

        self.stdout.write(
            self.style.SUCCESS(
                "Fichier traité (%(rows)d lignes) : %(updated)d produits mis à jour, "
                "%(not_found)d références introuvables, %(ambiguous)d correspondances multiples, "
                "%(invalid_cost)d coûts invalides, %(missing_reference)d références vides."
                % summary
            )
        )
        if errors:
            self.stdout.write(self.style.WARNING("Détails des lignes ignorées :"))
            for error in errors:
                self.stdout.write(f"- {error}")
